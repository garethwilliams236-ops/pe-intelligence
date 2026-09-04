#!/usr/bin/env python3
"""
WP2 :: load the Ardent master investor list into pe-intelligence.

Reads the master workbook and emits an idempotent SQL seed. Nothing is
written to the database directly, and the generated SQL is not committed —
it carries internal contact details, so it lands in out/ which is gitignored.

    python scripts/wp2_load_master_list.py "Master Investor List.xlsx"
    psql "$SUPABASE_DB_URL" -f out/wp2_master_investor_seed.sql

Design notes
------------
* The join key is public.normalise_name(), computed server-side, so this
  script never has to stay in sync with the SQL function.
* Identity is anchored on company_identifiers(scheme='ardent_master_list'),
  which already carries a unique constraint — so re-running is safe.
* Attribute data from this list is 2017-2019 vintage and is loaded as
  *claims*, not as accepted facts. Only names, aliases, websites and the
  mandate box land in typed columns, and the mandate box is stamped with
  its vintage so the ranking engine can discount it.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

SOURCE_CODE = "ardent_master_list"
LIST_AS_AT = "2026-02-09"  # workbook "26 2 09"; actively maintained

# --------------------------------------------------------------------------
# Name normalisation (replica of public.normalise_name, used only for
# in-memory merging; the database recomputes its own key on load)
# --------------------------------------------------------------------------
_SUFFIX = re.compile(
    r"\b(limited|ltd|plc|llp|lp|holdings?|group|company|co|incorporated|inc|"
    r"gmbh|ag|kg|kgaa|se|nv|bv|sa|sas|sarl|srl|spa|ab|as|a/s|aps|oy|oyj|"
    r"sl|sau|cvba|scrl|ug|ohg|gbr|eurl|snc)\b\.?",
    re.I,
)


def name_key(value) -> str | None:
    if value is None:
        return None
    text = _SUFFIX.sub(" ", str(value).strip().lower())
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text or None


def clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text.lower() in ("", "n/a", "na", "#n/a", "tbc", "none", "-", "0"):
        return None
    return re.sub(r"\s+", " ", text)


def sql(value) -> str:
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def num(value) -> str:
    return "null" if value is None else str(value)


# --------------------------------------------------------------------------
# Country resolution
# --------------------------------------------------------------------------
COUNTRY_BY_NAME = {
    "united kingdom": "GB", "uk": "GB", "england": "GB", "scotland": "GB",
    "wales": "GB", "northern ireland": "GB", "great britain": "GB",
    "ireland": "IE", "republic of ireland": "IE",
    "germany": "DE", "austria": "AT", "switzerland": "CH",
    "france": "FR", "the netherlands": "NL", "netherlands": "NL",
    "belgium": "BE", "luxembourg": "LU",
    "sweden": "SE", "norway": "NO", "denmark": "DK", "finland": "FI",
    "spain": "ES", "portugal": "PT", "italy": "IT",
    "united states": "US", "usa": "US", "canada": "CA",
}

CITY_BY_NAME = {
    "london": "GB", "manchester": "GB", "birmingham": "GB", "edinburgh": "GB",
    "leeds": "GB", "bristol": "GB", "cambridge": "GB", "oxford": "GB",
    "dublin": "IE", "paris": "FR", "lyon": "FR", "berlin": "DE",
    "munich": "DE", "münchen": "DE", "frankfurt": "DE", "hamburg": "DE",
    "vienna": "AT", "zurich": "CH", "zürich": "CH", "geneva": "CH",
    "amsterdam": "NL", "rotterdam": "NL", "brussels": "BE",
    "luxembourg": "LU", "stockholm": "SE", "oslo": "NO", "copenhagen": "DK",
    "helsinki": "FI", "madrid": "ES", "barcelona": "ES", "lisbon": "PT",
    "milan": "IT", "milano": "IT", "rome": "IT", "new york": "US",
    "san francisco": "US", "boston": "US", "chicago": "US", "palo alto": "US",
    "menlo park": "US", "los angeles": "US", "seattle": "US", "austin": "US",
    "toronto": "CA", "tel aviv": "IL", "singapore": "SG", "hong kong": "HK",
    "tokyo": "JP", "budapest": "HU", "warsaw": "PL", "prague": "CZ",
}


def resolve_country(location: str | None) -> str | None:
    if not location:
        return None
    parts = [p.strip().lower() for p in re.split(r"[,/]", location) if p.strip()]
    for part in reversed(parts):
        if part in COUNTRY_BY_NAME:
            return COUNTRY_BY_NAME[part]
    for part in parts:
        if part in CITY_BY_NAME:
            return CITY_BY_NAME[part]
    return None


# --------------------------------------------------------------------------
# Strategy and money parsing
# --------------------------------------------------------------------------
STRATEGY_RULES = [
    (r"private equity|buy.?out|lbo|mbo|mbi", "buyout"),
    (r"growth (capital|equity)|development capital|expansion", "growth"),
    (r"venture|vc\b|seed|angel|early stage", "venture"),
    (r"mezzanine", "mezzanine"),
    (r"debt|credit|lending", "credit"),
    (r"family office", "family_office"),
    (r"fund of funds", "fund_of_funds"),
    (r"secondar", "secondaries"),
    (r"infrastructure", "infrastructure"),
    (r"real estate|property", "real_estate"),
    (r"sovereign", "sovereign_wealth"),
    (r"pension", "pension"),
    (r"\bvct\b", "vct"),
    (r"\beis\b", "eis"),
    (r"special situation|distress|turnaround", "special_situations"),
]

FLAG_STRATEGY = {
    "LBO": "buyout", "MBO": "buyout", "VC": "venture",
    "SEED": "venture", "DC": "growth", "FO": "family_office",
}

CURRENCY_SYMBOL = {"£": "GBP", "€": "EUR", "$": "USD"}


def parse_strategies(*texts) -> list[str]:
    found: list[str] = []
    blob = " ".join(t for t in texts if t).lower()
    for pattern, strategy in STRATEGY_RULES:
        if re.search(pattern, blob) and strategy not in found:
            found.append(strategy)
    return found


def parse_money(value) -> tuple[float | None, str | None]:
    """Return (millions, currency). Source figures are stated in millions."""
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        return (float(value), None) if value else (None, None)
    text = str(value).strip()
    if not text or text.lower() in ("n/a", "#n/a", "tbc", "-"):
        return None, None
    currency = next((c for sym, c in CURRENCY_SYMBOL.items() if sym in text), None)
    match = re.search(r"(\d+(?:[.,]\d+)?)", text.replace(",", ""))
    if not match:
        return None, currency
    amount = float(match.group(1))
    if re.search(r"\bbn\b|billion", text, re.I):
        amount *= 1000
    if re.search(r"\bk\b|thousand", text, re.I):
        amount /= 1000
    return amount, currency


SENIORITY_RULES = [
    (r"managing partner|senior partner", "managing_partner"),
    (r"\bpartner\b", "partner"),
    (r"principal", "principal"),
    (r"managing director|\bmd\b|director", "director"),
    (r"investment manager", "investment_manager"),
    (r"associate", "associate"),
    (r"analyst", "analyst"),
    (r"operating partner", "operating_partner"),
    (r"chairman|chair\b", "chair"),
    (r"chief executive|\bceo\b|founder", "ceo"),
    (r"chief financial|\bcfo\b", "cfo"),
]


def parse_seniority(title: str | None) -> str:
    if not title:
        return "other"
    for pattern, seniority in SENIORITY_RULES:
        if re.search(pattern, title, re.I):
            return seniority
    return "other"


def domain_of(url: str | None) -> str | None:
    if not url:
        return None
    match = re.search(r"https?://(?:www\.)?([^/\s]+)", url.strip(), re.I)
    if match:
        return match.group(1).lower().rstrip(".")
    match = re.match(r"(?:www\.)?([a-z0-9.-]+\.[a-z]{2,})", url.strip(), re.I)
    return match.group(1).lower() if match else None


# --------------------------------------------------------------------------
# Sheet readers
# --------------------------------------------------------------------------
def read_sheet(workbook, sheet: str, header_row: int, name_col: int = 0):
    """Yield (name, {column: value}) using first-occurrence header names."""
    rows = list(workbook[sheet].iter_rows(min_row=header_row, values_only=True))
    header = [("" if h is None else str(h).strip()) for h in rows[0]]
    index: dict[str, int] = {}
    for position, column in enumerate(header):
        if column and column not in index:
            index[column] = position
    for row in rows[1:]:
        if name_col >= len(row):
            continue
        name = clean(row[name_col])
        if not name or name.lower() in ("name", "account name"):
            continue
        record = {col: (row[pos] if pos < len(row) else None) for col, pos in index.items()}
        yield name, record, row


class Investor:
    __slots__ = ("key", "names", "website", "country", "location", "description",
                 "strategies", "min_cheque", "max_cheque", "currency", "geography",
                 "sector_text", "fund_size", "deals", "contacts", "sheets")

    def __init__(self, key: str):
        self.key = key
        self.names: list[str] = []
        self.website = self.country = self.location = self.description = None
        self.strategies: list[str] = []
        self.min_cheque = self.max_cheque = self.currency = None
        self.geography = self.sector_text = self.fund_size = None
        self.deals: list[str] = []
        self.contacts: list[tuple[str, str | None, str | None]] = []
        self.sheets: set[str] = set()

    def add_name(self, name: str):
        if name not in self.names:
            self.names.append(name)

    @property
    def legal_name(self) -> str:
        # prefer the longest spelling — usually the most complete one
        return max(self.names, key=lambda n: (len(n), n))

    def fill(self, **kwargs):
        for field, value in kwargs.items():
            if value is not None and getattr(self, field) is None:
                setattr(self, field, value)


def build(path: Path) -> dict[str, Investor]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    investors: dict[str, Investor] = {}

    def get(key: str) -> Investor:
        return investors.setdefault(key, Investor(key))

    # --- Investor info: the only sheet carrying websites -------------------
    for name, rec, _ in read_sheet(workbook, "Investor info", 1):
        key = name_key(name)
        if not key:
            continue
        inv = get(key)
        inv.add_name(name)
        inv.sheets.add("Investor info")
        location = clean(rec.get("Location"))
        inv.fill(
            website=clean(rec.get("Website")),
            location=location,
            country=resolve_country(location),
            description=clean(rec.get("Full Description")),
            sector_text=clean(rec.get("Industry Groups")),
        )
        for strategy in parse_strategies(clean(rec.get("Investor Type")),
                                         clean(rec.get("Investment Stage"))):
            if strategy not in inv.strategies:
                inv.strategies.append(strategy)

    # --- InvestorBase Data: descriptions, cheque sizes, contacts -----------
    for name, rec, _ in read_sheet(workbook, "InvestorBase Data", 3):
        key = name_key(name)
        if not key:
            continue
        inv = get(key)
        inv.add_name(name)
        inv.sheets.add("InvestorBase Data")
        location = clean(rec.get("Location"))
        inv.fill(
            description=clean(rec.get("Business Description")),
            location=location,
            country=resolve_country(location),
            geography=clean(rec.get("Investment Geography")),
            sector_text=clean(rec.get("Sector Vertical")),
            fund_size=clean(rec.get("Latest Fund Size")),
        )
        low, _ = parse_money(rec.get("Min investment (£m)"))
        high, _ = parse_money(rec.get("Max investment (£m)"))
        if low is not None and inv.min_cheque is None:
            inv.min_cheque, inv.currency = low, inv.currency or "GBP"
        if high is not None and inv.max_cheque is None:
            inv.max_cheque, inv.currency = high, inv.currency or "GBP"
        for strategy in parse_strategies(clean(rec.get("Investment type")),
                                         clean(rec.get("Investment Stage"))):
            if strategy not in inv.strategies:
                inv.strategies.append(strategy)
        for n in (1, 2, 3, 4):
            person = clean(rec.get(f"Contact {n}"))
            if person:
                inv.contacts.append((person,
                                     clean(rec.get(f"Contact {n} title")),
                                     clean(rec.get(f"Contact {n} email"))))

    # --- Standard VCs: best cheque-size and deal coverage ------------------
    rows = list(workbook["Standard VCs"].iter_rows(min_row=9, values_only=True))
    for row in rows:
        name = clean(row[1]) if len(row) > 1 else None
        key = name_key(name)
        if not key:
            continue
        inv = get(key)
        inv.add_name(name)
        inv.sheets.add("Standard VCs")
        office = clean(row[2]) if len(row) > 2 else None
        inv.fill(location=office, country=resolve_country(office),
                 sector_text=clean(row[16]) if len(row) > 16 else None,
                 fund_size=clean(row[17]) if len(row) > 17 else None,
                 geography=clean(row[21]) if len(row) > 21 else None)
        for position, flag in [(3, "LP"), (4, "FO"), (5, "H"), (6, "LBO"),
                               (7, "MBO"), (8, "DC"), (9, "VC"), (10, "SEED")]:
            if position < len(row) and clean(row[position]):
                strategy = FLAG_STRATEGY.get(flag)
                if strategy and strategy not in inv.strategies:
                    inv.strategies.append(strategy)
        # Standard VCs states currency explicitly, so it wins over InvestorBase
        low, low_ccy = parse_money(row[19]) if len(row) > 19 else (None, None)
        high, high_ccy = parse_money(row[20]) if len(row) > 20 else (None, None)
        currency = low_ccy or high_ccy
        if currency and (low is not None or high is not None):
            inv.min_cheque, inv.max_cheque, inv.currency = low, high, currency
        for position in (22, 23, 24):
            deal = clean(row[position]) if len(row) > position else None
            if deal and deal not in inv.deals:
                inv.deals.append(deal)
        for position in (11, 12, 13):
            person = clean(row[position]) if len(row) > position else None
            if person:
                inv.contacts.append((person, None, None))

    # --- Combined Data: names and aliases only ----------------------------
    for name, _rec, _row in read_sheet(workbook, "Combined Data", 6):
        key = name_key(name)
        if key:
            inv = get(key)
            inv.add_name(name)
            inv.sheets.add("Combined Data")

    # A handful of source rows have the cheque range entered the wrong way
    # round (min €500m / max €6m). Swap rather than drop — the range is
    # still the useful fact — but count them so the mis-entry is visible.
    global INVERTED_CHEQUES
    INVERTED_CHEQUES = 0
    for inv in investors.values():
        if (inv.min_cheque is not None and inv.max_cheque is not None
                and inv.min_cheque > inv.max_cheque):
            inv.min_cheque, inv.max_cheque = inv.max_cheque, inv.min_cheque
            INVERTED_CHEQUES += 1

    return investors


INVERTED_CHEQUES = 0


# --------------------------------------------------------------------------
# SQL emission
# --------------------------------------------------------------------------
def emit(investors: dict[str, Investor], out: Path) -> dict[str, int]:
    people: dict[tuple[str, str], tuple[str, str | None, str | None]] = {}
    for inv in investors.values():
        for name, title, email in inv.contacts:
            if len(name) > 2 and not re.match(r"^[\d\W]+$", name):
                people[(inv.key, name_key(name) or name)] = (name, title, email)

    lines: list[str] = []
    add = lines.append

    add("-- Generated by scripts/wp2_load_master_list.py — do not edit by hand.")
    add("-- Contains internal contact details: keep out of version control.")
    add("begin;")
    add("")
    add("create temp table _stg_inv (")
    add("  src_key text primary key, legal_name text not null, website text, domain text,")
    add("  country_code char(2), location text, description text, geography text,")
    add("  sector_text text, fund_size text, min_cheque numeric, max_cheque numeric,")
    add("  currency char(3)) on commit drop;")
    add("")

    values = []
    for inv in sorted(investors.values(), key=lambda i: i.key):
        values.append(
            "  (" + ", ".join([
                sql(inv.key), sql(inv.legal_name), sql(inv.website),
                sql(domain_of(inv.website)), sql(inv.country), sql(inv.location),
                sql(inv.description), sql(inv.geography), sql(inv.sector_text),
                sql(inv.fund_size), num(inv.min_cheque), num(inv.max_cheque),
                sql(inv.currency),
            ]) + ")"
        )
    add("insert into _stg_inv values")
    add(",\n".join(values) + ";")
    add("")

    add("create temp table _stg_alias (src_key text, alias text) on commit drop;")
    alias_rows = [f"  ({sql(inv.key)}, {sql(name)})"
                  for inv in sorted(investors.values(), key=lambda i: i.key)
                  for name in inv.names[1:]]
    if alias_rows:
        add("insert into _stg_alias values")
        add(",\n".join(alias_rows) + ";")
    add("")

    add("create temp table _stg_strategy (src_key text, strategy text) on commit drop;")
    strategy_rows = [f"  ({sql(inv.key)}, {sql(s)})"
                     for inv in sorted(investors.values(), key=lambda i: i.key)
                     for s in inv.strategies]
    if strategy_rows:
        add("insert into _stg_strategy values")
        add(",\n".join(strategy_rows) + ";")
    add("")

    add("create temp table _stg_person (src_key text, full_name text, title text, email text) on commit drop;")
    person_rows = [f"  ({sql(src)}, {sql(name)}, {sql(title)}, {sql(email)})"
                   for (src, _), (name, title, email) in sorted(people.items())]
    if person_rows:
        add("insert into _stg_person values")
        add(",\n".join(person_rows) + ";")
    add("")

    add("create temp table _stg_claim (src_key text, attribute text, value_text text) on commit drop;")
    claim_rows = []
    for inv in sorted(investors.values(), key=lambda i: i.key):
        for deal in inv.deals:
            claim_rows.append(f"  ({sql(inv.key)}, 'relevant_deal', {sql(deal)})")
        if inv.sector_text:
            claim_rows.append(f"  ({sql(inv.key)}, 'stated_sector_text', {sql(inv.sector_text)})")
        if inv.geography:
            claim_rows.append(f"  ({sql(inv.key)}, 'stated_geography_text', {sql(inv.geography)})")
        if inv.fund_size:
            claim_rows.append(f"  ({sql(inv.key)}, 'stated_fund_size_text', {sql(inv.fund_size)})")
    if claim_rows:
        add("insert into _stg_claim values")
        add(",\n".join(claim_rows) + ";")
    add("")

    add(f"""
-- 1. companies -------------------------------------------------------------
insert into public.companies (legal_name, company_types, country_code, city,
                              website, website_domain, description, confidence)
select distinct on (public.normalise_name(s.legal_name))
       s.legal_name, '{{sponsor}}'::public.company_type[], s.country_code, s.location,
       s.website, s.domain, s.description, 0.600
from _stg_inv s
where public.normalise_name(s.legal_name) is not null
  and not exists (
    select 1 from public.companies c
    where c.name_key = public.normalise_name(s.legal_name))
order by public.normalise_name(s.legal_name), s.legal_name;

-- 2. anchor identity so re-running this file is a no-op --------------------
insert into public.company_identifiers (company_id, scheme, value)
select c.id, '{SOURCE_CODE}', s.src_key
from _stg_inv s
join public.companies c on c.name_key = public.normalise_name(s.legal_name)
on conflict (scheme, value) do nothing;

-- 3. aliases ---------------------------------------------------------------
insert into public.company_aliases (company_id, alias, alias_type)
select ci.company_id, a.alias, 'former'
from _stg_alias a
join public.company_identifiers ci on ci.scheme = '{SOURCE_CODE}' and ci.value = a.src_key
where not exists (
  select 1 from public.company_aliases ex
  where ex.company_id = ci.company_id
    and ex.alias_key = public.normalise_name(a.alias));

-- 4. investors, with the mandate box stamped with its vintage --------------
insert into public.investors (company_id, hq_country_code, min_cheque_raw, max_cheque_raw,
                              cheque_currency, min_equity_cheque_gbp, max_equity_cheque_gbp,
                              mandate_box_as_at, mandate_box_source_id, notes)
select ci.company_id, s.country_code,
       s.min_cheque * 1e6, s.max_cheque * 1e6, s.currency,
       case when s.currency = 'GBP' then s.min_cheque * 1e6 end,
       case when s.currency = 'GBP' then s.max_cheque * 1e6 end,
       '{LIST_AS_AT}'::date,
       (select id from public.sources where code = '{SOURCE_CODE}'),
       nullif(concat_ws(' | ',
         nullif('geography: ' || s.geography, 'geography: '),
         nullif('fund size: ' || s.fund_size, 'fund size: ')), '')
from _stg_inv s
join public.company_identifiers ci on ci.scheme = '{SOURCE_CODE}' and ci.value = s.src_key
on conflict (company_id) do nothing;

-- 5. strategies ------------------------------------------------------------
insert into public.investor_strategies (company_id, strategy)
select distinct ci.company_id, st.strategy::public.investor_strategy
from _stg_strategy st
join public.company_identifiers ci on ci.scheme = '{SOURCE_CODE}' and ci.value = st.src_key
join public.investors i on i.company_id = ci.company_id
on conflict (company_id, strategy) do nothing;

-- 6. contacts as people + roles -------------------------------------------
insert into public.people (full_name, confidence)
select distinct on (public.normalise_name(p.full_name)) p.full_name, 0.500
from _stg_person p
where public.normalise_name(p.full_name) is not null
  and not exists (
    select 1 from public.people ex
    where ex.name_key = public.normalise_name(p.full_name))
order by public.normalise_name(p.full_name), p.full_name;

insert into public.person_roles (person_id, company_id, title, seniority, source_id)
select distinct pe.id, ci.company_id, p.title,
       case
         when p.title ~* 'managing partner|senior partner' then 'managing_partner'
         when p.title ~* 'operating partner'               then 'operating_partner'
         when p.title ~* 'partner'                         then 'partner'
         when p.title ~* 'principal'                       then 'principal'
         when p.title ~* 'managing director|director'      then 'director'
         when p.title ~* 'investment manager'              then 'investment_manager'
         when p.title ~* 'associate'                       then 'associate'
         when p.title ~* 'analyst'                         then 'analyst'
         when p.title ~* 'chairman|chair'                  then 'chair'
         when p.title ~* 'chief executive|ceo|founder'     then 'ceo'
         when p.title ~* 'chief financial|cfo'             then 'cfo'
         else 'other'
       end::public.person_seniority,
       (select id from public.sources where code = '{SOURCE_CODE}')
from _stg_person p
join public.company_identifiers ci on ci.scheme = '{SOURCE_CODE}' and ci.value = p.src_key
join public.people pe on pe.name_key = public.normalise_name(p.full_name)
where not exists (
  select 1 from public.person_roles ex
  where ex.person_id = pe.id and ex.company_id = ci.company_id);

-- 7. everything else lands as a candidate claim, not as fact ---------------
insert into public.claims (subject_table, subject_id, subject_key, attribute,
                           value_text, status, confidence, licence_class,
                           extracted_by, extraction_version)
select 'investors', ci.company_id, cl.src_key, cl.attribute, cl.value_text,
       'candidate', 0.400, 'confidential', 'wp2_load_master_list', '1'
from _stg_claim cl
join public.company_identifiers ci on ci.scheme = '{SOURCE_CODE}' and ci.value = cl.src_key
where not exists (
  select 1 from public.claims ex
  where ex.subject_table = 'investors' and ex.subject_id = ci.company_id
    and ex.attribute = cl.attribute and ex.value_text = cl.value_text);

-- 8. seed the crawler off the websites we now hold -------------------------
insert into public.crawl_domains (domain)
select distinct s.domain from _stg_inv s where s.domain is not null
on conflict (domain) do nothing;

insert into public.crawl_targets (source_id, company_id, label, start_url, domain,
                                  target_kind, frequency_hours, is_enabled)
select (select id from public.sources where code = 'sponsor_site'),
       ci.company_id, s.legal_name || ' — site', s.website, s.domain,
       'sponsor_root', 720, false
from _stg_inv s
join public.company_identifiers ci on ci.scheme = '{SOURCE_CODE}' and ci.value = s.src_key
where s.website is not null and s.domain is not null
  and not exists (
    select 1 from public.crawl_targets ct
    where ct.company_id = ci.company_id and ct.target_kind = 'sponsor_root');

commit;
""")

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(lines), encoding="utf-8")

    return {
        "investors": len(investors),
        "aliases": len(alias_rows),
        "with_website": sum(1 for i in investors.values() if i.website),
        "with_country": sum(1 for i in investors.values() if i.country),
        "with_cheque": sum(1 for i in investors.values() if i.min_cheque or i.max_cheque),
        "with_strategy": sum(1 for i in investors.values() if i.strategies),
        "people": len(people),
        "claims": len(claim_rows),
        "cheque_swapped": INVERTED_CHEQUES,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("-o", "--out", type=Path,
                        default=Path("out/wp2_master_investor_seed.sql"))
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    investors = build(args.workbook)
    stats = emit(investors, args.out)

    print(f"wrote {args.out}")
    for label, count in stats.items():
        print(f"  {label:14s} {count:5d}")

    by_country: dict[str, int] = defaultdict(int)
    for inv in investors.values():
        by_country[inv.country or "unknown"] += 1
    in_scope = {"GB", "IE", "DE", "AT", "CH", "FR", "NL", "BE", "LU",
                "SE", "NO", "DK", "FI", "ES", "PT", "IT"}
    scoped = sum(v for k, v in by_country.items() if k in in_scope)
    print(f"  in UK/W-Europe {scoped:5d}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
